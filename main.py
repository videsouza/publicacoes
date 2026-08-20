import sqlite3
import io
import pandas as pd
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import List
from ortools.sat.python import cp_model

app = FastAPI(title="Sistema Base 2.0")

# ============================================================================
# INICIALIZAÇÃO DO BANCO DE DADOS
# ============================================================================
def inicializar_banco():
    conexao = sqlite3.connect("banco_sistema.db")
    cursor = conexao.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS turmas (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS disciplinas (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS professores (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT)")
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS matrizes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            turma_id INTEGER,
            disciplina_id INTEGER,
            professor_id INTEGER,
            aulas INTEGER
        )
    ''')
    # NOVA TABELA PARA RESTRIÇÕES
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS disponibilidades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            professor_id INTEGER,
            dia INTEGER,
            periodo INTEGER,
            tipo TEXT 
        )
    ''')
    conexao.commit()
    conexao.close()

inicializar_banco()

# ============================================================================
# ROTAS VISUAIS E MODELOS
# ============================================================================
@app.get("/")
def renderizar_painel():
    return FileResponse("index.html")

class TurmaBase(BaseModel): nome: str
class DisciplinaBase(BaseModel): nome: str
class ProfessorBase(BaseModel): nome: str
class MatrizBase(BaseModel):
    turma_id: int
    disciplina_id: int
    professor_id: int
    aulas: int
class ConfigGrade(BaseModel): prioridades: List[str] = []
class ItemGrade(BaseModel):
    turma: str
    disciplina: str
    professor: str
    dia: int
    periodo: int

# (As rotas básicas de CRUD de turmas, disciplinas, professores e matrizes permanecem iguais)
@app.get("/api/turmas")
def listar_turmas():
    conexao = sqlite3.connect("banco_sistema.db")
    linhas = conexao.execute("SELECT id, nome FROM turmas").fetchall()
    conexao.close()
    return [{"id": l[0], "nome": l[1]} for l in linhas]

@app.get("/api/disciplinas")
def listar_disciplinas():
    conexao = sqlite3.connect("banco_sistema.db")
    linhas = conexao.execute("SELECT id, nome FROM disciplinas").fetchall()
    conexao.close()
    return [{"id": l[0], "nome": l[1]} for l in linhas]

@app.get("/api/professores")
def listar_professores():
    conexao = sqlite3.connect("banco_sistema.db")
    linhas = conexao.execute("SELECT id, nome FROM professores").fetchall()
    conexao.close()
    return [{"id": l[0], "nome": l[1]} for l in linhas]

@app.get("/api/matrizes")
def listar_matrizes():
    conexao = sqlite3.connect("banco_sistema.db")
    linhas = conexao.execute('''
        SELECT m.id, t.nome, d.nome, p.nome, m.aulas 
        FROM matrizes m JOIN turmas t ON m.turma_id = t.id JOIN disciplinas d ON m.disciplina_id = d.id JOIN professores p ON m.professor_id = p.id
    ''').fetchall()
    conexao.close()
    return [{"id": l[0], "turma": l[1], "disciplina": l[2], "professor": l[3], "aulas": l[4]} for l in linhas]

# ============================================================================
# ROTA DE IMPORTAÇÃO (AGORA COM 3 PLANILHAS)
# ============================================================================
@app.post("/api/upload-cadastros")
async def upload_cadastros(file: UploadFile = File(...)):
    conteudo = await file.read()
    
    try:
        df_prof = pd.read_excel(io.BytesIO(conteudo), sheet_name=0)
        df_aulas = pd.read_excel(io.BytesIO(conteudo), sheet_name=1)
        # Tenta ler a Planilha3 (se não existir, não quebra o código)
        try:
            df_disp = pd.read_excel(io.BytesIO(conteudo), sheet_name=2)
        except:
            df_disp = pd.DataFrame()
    except Exception as e:
        return {"erro": "Certifique-se de que o arquivo Excel possui as abas Planilha1 e Planilha2."}

    df_prof.rename(columns=lambda x: str(x).strip().upper(), inplace=True)
    df_aulas.rename(columns=lambda x: str(x).strip().upper(), inplace=True)
    
    conexao = sqlite3.connect("banco_sistema.db")
    cursor = conexao.cursor()
    cursor.execute("DELETE FROM turmas"); cursor.execute("DELETE FROM disciplinas")
    cursor.execute("DELETE FROM professores"); cursor.execute("DELETE FROM matrizes")
    cursor.execute("DELETE FROM disponibilidades")
    
    # Processamento Padrão de Turmas e Matrizes
    turmas_set = set(); professores_set = set(); disciplinas_set = set(col for col in df_prof.columns if col != 'TURMA')
    for index, row in df_prof.iterrows():
        turma = str(row['TURMA']).strip()
        if pd.isna(turma) or not turma or turma.lower() == 'nan': continue
        turmas_set.add(turma)
        for disc in disciplinas_set:
            prof = str(row.get(disc, '')).strip()
            if prof and prof.lower() != 'nan': professores_set.add(prof)

    mapa_turmas = {t: cursor.execute("INSERT INTO turmas (nome) VALUES (?)", (t,)).lastrowid for t in turmas_set}
    mapa_disc = {d: cursor.execute("INSERT INTO disciplinas (nome) VALUES (?)", (d,)).lastrowid for d in disciplinas_set}
    mapa_prof = {p: cursor.execute("INSERT INTO professores (nome) VALUES (?)", (p,)).lastrowid for p in professores_set}

    vinculos = 0
    for index, row_prof in df_prof.iterrows():
        turma_nome = str(row_prof['TURMA']).strip()
        if turma_nome not in mapa_turmas: continue
        row_aulas = df_aulas[df_aulas['TURMA'].astype(str).str.strip() == turma_nome]
        if row_aulas.empty: continue
        row_aulas = row_aulas.iloc[0] 
        for disc_nome in disciplinas_set:
            prof_nome = str(row_prof.get(disc_nome, '')).strip()
            try: qtd_aulas = int(float(row_aulas.get(disc_nome, 0)))
            except: qtd_aulas = 0
            if prof_nome and prof_nome.lower() != 'nan' and qtd_aulas > 0:
                cursor.execute("INSERT INTO matrizes (turma_id, disciplina_id, professor_id, aulas) VALUES (?, ?, ?, ?)", 
                               (mapa_turmas[turma_nome], mapa_disc[disc_nome], mapa_prof[prof_nome], qtd_aulas))
                vinculos += 1

    # ---------------------------------------------------------
    # PROCESSAMENTO DA PLANILHA 3 (Restrições)
    # ---------------------------------------------------------
    if not df_disp.empty:
        df_disp.rename(columns=lambda x: str(x).strip().upper(), inplace=True)
        mapa_dias = {'SEG': 0, 'TER': 1, 'QUA': 2, 'QUI': 3, 'SEX': 4}
        
        for index, row in df_disp.iterrows():
            dia_str = str(row.get('DIA', '')).strip().upper()
            if dia_str not in mapa_dias: continue
            dia_idx = mapa_dias[dia_str]
            
            try: aula_idx = int(float(row.get('AULA', 0))) - 1 # Converte aula 1 para índice 0
            except: continue
            
            # Varre todas as colunas de professores
            for col in df_disp.columns:
                if col in ['DIA', 'AULA']: continue
                
                marca = str(row.get(col, '')).strip().upper()
                if marca in ['X', 'Y']:
                    prof_nome = col
                    if prof_nome in mapa_prof: # Só cadastra se o professor der aula na escola
                        cursor.execute("INSERT INTO disponibilidades (professor_id, dia, periodo, tipo) VALUES (?, ?, ?, ?)",
                                      (mapa_prof[prof_nome], dia_idx, aula_idx, marca))

    conexao.commit()
    conexao.close()
    return {"mensagem": "Automação concluída!", "resumo": {"turmas": len(turmas_set), "professores": len(professores_set), "vinculos": vinculos}}

# ============================================================================
# MOTOR DE GERAÇÃO E DIAGNÓSTICO (OR-TOOLS)
# ============================================================================
@app.post("/api/gerar-grade")
def gerar_grade_mestra(config: ConfigGrade):
    conexao = sqlite3.connect("banco_sistema.db")
    cursor = conexao.cursor()
    
    matriz = cursor.execute('''
        SELECT m.id, t.nome, d.nome, p.nome, m.aulas FROM matrizes m
        JOIN turmas t ON m.turma_id = t.id JOIN disciplinas d ON m.disciplina_id = d.id JOIN professores p ON m.professor_id = p.id
    ''').fetchall()
    
    restricoes_brutas = cursor.execute('''
        SELECT p.nome, d.dia, d.periodo, d.tipo FROM disponibilidades d
        JOIN professores p ON d.professor_id = p.id
    ''').fetchall()
    conexao.close()

    if not matriz: return {"erro": "A matriz curricular está vazia."}

    dias = range(5)
    periodos = range(6)
    professores_unicos = set(m[3] for m in matriz)
    
    # ------------------------------------------------------------------------
    # MÓDULO DE PRÉ-CHECAGEM (O RASTREADOR DE IMPASSES)
    # ------------------------------------------------------------------------
    # 1. Checa a carga horária vs bloqueios de cada professor
    for prof in professores_unicos:
        carga_total = sum(m[4] for m in matriz if m[3] == prof)
        bloqueios_x = sum(1 for r in restricoes_brutas if r[0] == prof and r[3] == 'X')
        horarios_livres = 30 - bloqueios_x
        
        if carga_total > horarios_livres:
            return {"erro": f"ERRO DE MATRIZ: O(a) professor(a) {prof} possui {carga_total} aulas alocadas, mas tem apenas {horarios_livres} horários disponíveis devido aos 'X' marcados na Planilha3."}

    # 2. Checa limite de carga horária por turma (Máx 30)
    turmas_unicas = set(m[1] for m in matriz)
    for turma in turmas_unicas:
        carga_turma = sum(m[4] for m in matriz if m[1] == turma)
        if carga_turma > 30:
            return {"erro": f"ERRO DE MATRIZ: A turma {turma} tem {carga_turma} aulas cadastradas, mas a semana só possui 30 horários físicos."}

    # ------------------------------------------------------------------------
    # CONFIGURAÇÃO DO ROBÔ (MODELO)
    # ------------------------------------------------------------------------
    modelo = cp_model.CpModel()
    grade = {}
    for m in matriz:
        m_id = m[0]
        for d in dias:
            for p in periodos:
                grade[(m_id, d, p)] = modelo.NewBoolVar(f'aula_{m_id}_d{d}_p{p}')

    # C1. Cumprir a Matriz
    for m in matriz:
        modelo.Add(sum(grade[(m[0], d, p)] for d in dias for p in periodos) == m[4])

    # C2. Colisões Básicas
    for d in dias:
        for p in periodos:
            for turma in turmas_unicas:
                modelo.AddAtMostOne([grade[(m[0], d, p)] for m in matriz if m[1] == turma])
            for prof in professores_unicos:
                modelo.AddAtMostOne([grade[(m[0], d, p)] for m in matriz if m[3] == prof])

    # C3. Fadiga
    for d in dias:
        for prof in professores_unicos:
            for turma in turmas_unicas:
                aulas = [grade[(m[0], d, p)] for m in matriz if m[3] == prof and m[1] == turma for p in periodos]
                if aulas: modelo.Add(sum(aulas) <= 2)

    # ------------------------------------------------------------------------
    # C4. REGRAS DA PLANILHA 3 (X e Y) E OTIMIZAÇÃO GERAL
    # ------------------------------------------------------------------------
    variaveis_recompensa = []
    
    # Aplica os bloqueios (X) e Preferências (Y)
    for res in restricoes_brutas:
        prof_nome, d, p, tipo = res
        aulas_naquele_momento = [grade[(m[0], d, p)] for m in matriz if m[3] == prof_nome]
        
        if aulas_naquele_momento:
            if tipo == 'X':
                # Regra de Aço: A soma de aulas neste exato horário para este professor DEVE ser ZERO.
                modelo.Add(sum(aulas_naquele_momento) == 0)
            elif tipo == 'Y':
                # Preferência: Se a inteligência conseguir botar o professor aqui, ela ganha 20 pontos.
                variaveis_recompensa.append(sum(aulas_naquele_momento) * 20)

    # Dobradinhas Prioritárias
    padrao_zero = [0, 0, 0, 0, 0, 0]
    padroes_um = [[1 if i == p else 0 for i in range(6)] for p in range(6)]
    padroes_dois = [[1 if i == p or i == p+1 else 0 for i in range(6)] for p in range(5)]
    lista_prioridades = [p.strip().upper() for p in config.prioridades]

    for m in matriz:
        m_id = m[0]
        nome_disciplina = str(m[2]).strip().upper()
        eh_prioridade = nome_disciplina in lista_prioridades
        
        for d in dias:
            aulas_no_dia = [grade[(m_id, d, p)] for p in periodos]
            modelo.AddAllowedAssignments(aulas_no_dia, [padrao_zero] + padroes_um + padroes_dois)
            
            if eh_prioridade:
                tem_duas = modelo.NewBoolVar(f'duas_{m_id}_d{d}')
                modelo.Add(sum(aulas_no_dia) == 2).OnlyEnforceIf(tem_duas)
                modelo.Add(sum(aulas_no_dia) < 2).OnlyEnforceIf(tem_duas.Not())
                # As dobradinhas valem 100 pontos, garantindo que a IA priorize a didática sobre as preferências 'Y' (20 pts)
                variaveis_recompensa.append(tem_duas * 100)

    if variaveis_recompensa:
        modelo.Maximize(sum(variaveis_recompensa))

    # ------------------------------------------------------------------------
    # PASSO D: RESOLUÇÃO E TRATAMENTO DE FALHA
    # ------------------------------------------------------------------------
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 60.0 
    solver.parameters.num_search_workers = 8 

    status = solver.Solve(modelo)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        resultado_grade = []
        for m in matriz:
            for d in dias:
                for p in periodos:
                    if solver.Value(grade[(m[0], d, p)]) == 1:
                        resultado_grade.append({"turma": m[1], "disciplina": m[2], "professor": m[3], "dia": d, "periodo": p})
        return {"mensagem": "Grade gerada!", "status": "sucesso", "grade": resultado_grade}
    
    else:
        # Se os dados passaram pela pré-checagem matemática mas o algoritmo falhou:
        return {"erro": "Impossível Gerar! Os horários solicitados formam um 'Nó Geométrico'. Tente remover algumas restrições com 'X' da Planilha 3, pois a IA não encontrou espaço físico para cruzar todas as turmas e professores simultaneamente."}

# ============================================================================
# ROTA DE EXPORTAÇÃO PARA EXCEL
# ============================================================================
@app.post("/api/exportar-grade")
def exportar_grade_excel(grade: List[ItemGrade]):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    wb = Workbook(); ws = wb.active; ws.title = "Grade_Mestra"
    turmas = sorted(list(set(item.turma for item in grade)))
    professores_unicos = sorted(list(set(item.professor for item in grade)))
    dias_nomes = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX"}
    mapa_aulas = {(item.dia, item.periodo, item.turma): {"texto": f"{item.disciplina}\n{item.professor}", "professor": item.professor} for item in grade}
    paleta_hex = ["FFB3BA", "FFDFBA", "FFFFBA", "BAFFC9", "BAE1FF", "D0E6A5", "FFCCB6", "F3B0C3", "C6DBDA", "FEE1E8", "FED7C3", "F6EAC2", "ECD5E3", "CBAACB", "FF968A", "8FCACA", "CCE2CB", "B6CFB6", "97C1A9", "FCB9AA"]
    mapa_cores = {prof: PatternFill(start_color=paleta_hex[i % len(paleta_hex)], end_color=paleta_hex[i % len(paleta_hex)], fill_type="solid") for i, prof in enumerate(professores_unicos)}
    borda_fina = Side(border_style="thin", color="000000"); borda_caixa = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina); alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True); fonte_negrito = Font(bold=True)
    ws.cell(row=1, column=1, value="DIA").alignment = alinhamento_centro; ws.cell(row=1, column=2, value="AULA").alignment = alinhamento_centro; ws.cell(row=1, column=1).font = fonte_negrito; ws.cell(row=1, column=2).font = fonte_negrito; ws.cell(row=1, column=1).border = borda_caixa; ws.cell(row=1, column=2).border = borda_caixa
    for col_idx, turma in enumerate(turmas, start=3):
        cel = ws.cell(row=1, column=col_idx, value=turma); cel.alignment = alinhamento_centro; cel.font = fonte_negrito; cel.border = borda_caixa
    linha_atual = 2
    for dia in range(5):
        linha_inicio_dia = linha_atual
        for periodo in range(6):
            cel_p = ws.cell(row=linha_atual, column=2, value=periodo + 1); cel_p.alignment = alinhamento_centro; cel_p.font = fonte_negrito; cel_p.border = borda_caixa
            for col_idx, turma in enumerate(turmas, start=3):
                cel_a = ws.cell(row=linha_atual, column=col_idx); cel_a.alignment = alinhamento_centro; cel_a.border = borda_caixa
                aula = mapa_aulas.get((dia, periodo, turma))
                if aula: cel_a.value = aula["texto"]; cel_a.fill = mapa_cores[aula["professor"]]
            linha_atual += 1
        ws.merge_cells(start_row=linha_inicio_dia, start_column=1, end_row=linha_atual - 1, end_column=1)
        cel_d = ws.cell(row=linha_inicio_dia, column=1, value=dias_nomes[dia]); cel_d.alignment = alinhamento_centro; cel_d.font = fonte_negrito
        for r in range(linha_inicio_dia, linha_atual): ws.cell(row=r, column=1).border = borda_caixa
    ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 8
    for col_idx in range(3, len(turmas) + 3): ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    for row_idx in range(2, linha_atual): ws.row_dimensions[row_idx].height = 35
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, headers={'Content-Disposition': 'attachment; filename="Grade_Escolar_Matriz.xlsx"'}, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
